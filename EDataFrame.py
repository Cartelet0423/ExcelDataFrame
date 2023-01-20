"""
pandas.DataFrameの拡張クラス
Excelとのデータ授受を容易に行うことができます。

DataFrame.push() :: Excelでデータ開きます。Excelウィンドウが開いていない場合はExcelを起動します。実行時に表示しているデータは上書きされる点に注意してください。
DataFrame.pull() :: Excelで表示しているデータをDataFrameとして受け取ります。Excelウィンドウが開いていない場合はExcelを起動します。
"""

import pandas as pd
from datetime import datetime
import win32com.client
import pywintypes
import numpy as np
from openpyxl.utils import get_column_letter


class DataFrame(pd.DataFrame):
    @property
    def _constructor(self):
        return DataFrame

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def push(self):
        self._connect()
        self.sheet.Cells.Clear()
        row, col = self.shape
        headers = [str(i) for i in self.columns]
        self.sheet.Range("A1", f"{get_column_letter(col)}1").NumberFormatLocal = "@"
        self.sheet.Range("A1", f"{get_column_letter(col)}1").Value = headers
        values = self._to_pydatetime(self.values)
        self.sheet.Range("A2", f"{get_column_letter(col)}{row+1}").Value = values
        return self

    def pull(self):
        self._connect()
        col, row = self._get_range()
        headers = self.sheet.Range("A1", f"{get_column_letter(col)}1").Value
        headers = headers[0] if isinstance(headers, tuple) else [headers]
        data = self._to_pydatetime(
            self.sheet.Range("A2", f"{get_column_letter(col)}{row+1}").Value
        )
        return DataFrame(data, columns=headers)

    def _get_range(self):
        return (
            self.sheet.Cells(1, self.sheet.Columns.Count).End(-4159).Column,
            self.sheet.Cells(self.sheet.Rows.Count, 1).End(-4162).Row - 1,
        )

    def _connect(self):
        try:
            self.excel = win32com.client.GetActiveObject("Excel.Application")
            if not self.excel.visible:
                self.excel.visible = True
                self.workbook = self.excel.Workbooks.Add()
            else:
                self.workbook = self.excel.Workbooks[0]
            self.sheet = self.workbook.Worksheets("Sheet1")
        except Exception as e:
            self.excel = win32com.client.Dispatch("Excel.Application")
            self.excel.visible = True
            self.workbook = self.excel.Workbooks.Add()
            self.sheet = self.workbook.Worksheets("Sheet1")

    def _to_pydatetime(self, x):
        def f(x):
            return (
                x.to_pydatetime()
                if isinstance(x, pd._libs.tslibs.timestamps.Timestamp)
                else datetime.fromisoformat(x.isoformat())
                if isinstance(x, pywintypes.TimeType)
                else x
            )

        return np.vectorize(f)(x)
